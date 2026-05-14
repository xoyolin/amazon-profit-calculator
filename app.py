import streamlit as st
import pandas as pd
import numpy as np
import re
import io
import os
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def load_mappings():
    mapping_file = "Mapping.xlsx"
    if not os.path.exists(mapping_file):
        st.error(f"❌ 未找到 {mapping_file} 文件！")
        st.stop()
    
    try:
        col_df = pd.read_excel(mapping_file, sheet_name="Columns")
        val_df = pd.read_excel(mapping_file, sheet_name="Values")
        disp_df = pd.read_excel(mapping_file, sheet_name="Display")
        month_df = pd.read_excel(mapping_file, sheet_name="Months") 
        
        builtin_col_map = dict(zip(col_df['Original'].astype(str).str.strip().str.lower(), col_df['Standard'].astype(str).str.strip()))
        builtin_val_map = dict(zip(val_df['Original'].astype(str).str.strip().str.lower(), val_df['Standard'].astype(str).str.strip()))
        type_mapping = dict(zip(disp_df['Key'].astype(str).str.strip(), disp_df['Display'].astype(str).str.strip()))
        
        month_mapping = dict(zip(month_df['Original'].astype(str).str.strip().str.lower(), month_df['Standard']))
        
        return builtin_col_map, builtin_val_map, type_mapping, month_mapping 
    except Exception as e:
        st.error(f"❌ 读取映射失败: {e}")
        st.stop()

BUILTIN_COL_MAP, BUILTIN_VAL_MAP, TYPE_MAPPING, MONTH_MAP = load_mappings()


def clean_and_parse_date(val):
    if pd.isna(val) or str(val).strip() == "": return pd.NaT
    vc = re.sub(r'\s+[A-Za-z]{3,4}$', '', str(val).strip()).strip()
    try:
        if re.search(r'\d{1,2}\.\d{1,2}\.\d{4}', vc): return pd.to_datetime(vc, dayfirst=True)
        elif re.search(r'\d{1,2}/\d{1,2}/\d{4}', vc): return pd.to_datetime(vc, dayfirst=False)
        else: return pd.to_datetime(vc)
    except: pass

    cs = re.sub(r'\s+', ' ', re.sub(r'[,\.\-]', ' ', vc.lower().strip()))
    parts = cs.split(' ')
    if len(parts) < 3: return pd.NaT
    try:
        year = int(parts[2])
        d, m_s = (int(parts[0]), parts[1]) if parts[0].isdigit() else (int(parts[1]), parts[0])
        m_s = m_s.replace('é','e').replace('ä','a')
        
        month = MONTH_MAP.get(m_s, 0)
        
        return datetime(year, month, d) if month > 0 else pd.NaT
    except: return pd.NaT

def safe_parse_number(val):
    if pd.isna(val) or isinstance(val, (int, float)): return val
    s = str(val).strip().replace(chr(8722), "-")
    if not s or not re.match(r'^-?[\d\s\.,]+$', s): return val
    vc = s.replace(' ', '').replace('\xa0', '')
    ld, lc = vc.rfind('.'), vc.rfind(',')
    try:
        if lc > ld and lc != -1: return float(vc.replace('.', '').replace(',', '.'))
        elif ld > lc and ld != -1: return float(vc.replace(',', ''))
        elif ld != -1 and lc == -1: return float(vc.replace('.', '')) if len(vc)-ld-1==3 else float(vc)
        elif lc != -1 and ld == -1: return float(vc.replace(',', '')) if len(vc)-lc-1==3 else float(vc.replace(',', '.'))
        return float(vc)
    except: return val

def map_amazon_type(row):
    t, d = str(row.get('type', '')).lower(), str(row.get('description', '')).strip()
    if t == 'service fee':
        if 'Coupon Redemption Fee' in d: return 'coupon redemption fee'
        return d.lower()
    if t in ('', 'nan', 'none', 'others'):
        if d.startswith('Save'): return 'coupon redemption fee'
        return 'others'
    if 'FBA Long-Term Storage Fee' in d: return 'fba long-term storage fee'
    if 'FBA Inventory Storage Fee' in d: return 'fba inventory storage fee'
    return t

st.set_page_config(page_title="亚马逊财务利润统计系统", layout="centered")
st.title("📊 亚马逊财务利润统计系统v1.0")
st.caption("已支持美国/加拿大/英国/德国/法国/意大利/西班牙/瑞典/荷兰/波兰/比利时/爱尔兰 12 个国家")

st.info(
    "🛡️ **数据安全承诺**：代码已在GitHub开源。您的所有报表数据仅在内存中进行即时核算，"
    "**绝不会被上传、收集或储存在任何服务器上**。页面刷新或关闭后数据立即销毁，请放心使用！"
)
st.markdown("---") 

with st.sidebar:
    st.header("☕ 赞助与支持")
    st.markdown("如果这个开源工具帮您节省了核算利润的时间，欢迎请作者喝一杯奶茶！您的支持是我持续更新的最大动力。")
    
    if os.path.exists("pay.png"):
        st.image("pay.png", caption="打开微信扫一扫")

if 'raw_df' not in st.session_state: st.session_state.raw_df = None
if 'sku_list' not in st.session_state: st.session_state.sku_list = None

uploaded_report = st.file_uploader("请上传亚马逊日期范围报告 (CSV)", type=["csv"], key="report_up")

if uploaded_report:
    try:
        raw_bytes = uploaded_report.getvalue()
        content = raw_bytes.decode('utf-8', errors='ignore')
        lines = content.splitlines()
        
        header_idx, detected_sep = 0, ','
        for i, line in enumerate(lines[:30]):
            clean_line = line.replace('"', '')
            comma_c, semi_c, tab_c = clean_line.count(','), clean_line.count(';'), clean_line.count('\t')
            c_max = max(comma_c, semi_c, tab_c)
            if c_max > 15: 
                current_sep = ','
                if c_max == semi_c: current_sep = ';'
                elif c_max == tab_c: current_sep = '\t'
                fields = clean_line.split(current_sep)
                if sum(1 for f in fields if len(f.strip()) > 0) > 10:
                    header_idx, detected_sep = i, current_sep
                    break
                    
        uploaded_report.seek(0)
        df = pd.read_csv(uploaded_report, encoding='utf-8', sep=detected_sep, engine='python', skiprows=header_idx)
        
        original_cols = list(df.columns)
        if len(original_cols) > 0: original_cols[0] = 'date'
        df.columns = [str(c).strip().lower().replace(":","").replace('"','').replace('：','') for c in original_cols]
        

        df.rename(columns=BUILTIN_COL_MAP, inplace=True)
        
        exclude = {'date','sku','type','description','order id','marketplace','fulfillment','settlement id','abrechnungsnummer'}
        for c in df.columns:
            if c.lower() not in exclude:
                df[c] = df[c].apply(safe_parse_number)
                
        if 'type' in df.columns:

            df['type'] = df['type'].astype(str).str.strip().str.lower().apply(lambda x: BUILTIN_VAL_MAP.get(x, x))
            df['type'] = df.apply(map_amazon_type, axis=1)
            df['type'] = df['type'].apply(lambda x: BUILTIN_VAL_MAP.get(x, x))
        else:
            st.error("警告：找不到 Type 列。")

        calc_cols = ['total','tax_sales','tax_shipping','tax_giftwrap','tax_promo','tax_withheld']
        for c in calc_cols:
            if c not in df.columns: df[c] = 0.0
            else: df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)
            
        df['net total'] = df['total'] - (df['tax_sales'] + df['tax_shipping'] + df['tax_giftwrap'] + df['tax_promo'] + df['tax_withheld'])
        if 'date' in df.columns: df['date'] = df['date'].apply(clean_and_parse_date)
        if 'sku' in df.columns: df['sku'] = df['sku'].astype(str).str.strip()
        
        st.session_state.raw_df = df
        st.success("✅ 报告解析成功！映射已生效。")

    except Exception as e: st.error(f"解析失败: {e}")

if st.session_state.raw_df is not None:
    st.markdown("---")
    if 'type' in st.session_state.raw_df.columns:
        df_f = st.session_state.raw_df[st.session_state.raw_df['type'].isin(['order', 'refund'])]
        if 'sku' in df_f.columns and 'description' in df_f.columns:
            u_skus = df_f[['sku', 'description']].dropna(subset=['sku']).drop_duplicates(subset=['sku'])
            if not u_skus.empty:
                tmpl = pd.DataFrame({'产品描述': u_skus['description'], 'SKU': u_skus['sku'], '产品和头程成本': 0.0})
                buf = io.BytesIO()
                with pd.ExcelWriter(buf, engine='openpyxl') as w: tmpl.to_excel(w, index=False)
                st.download_button("⬇️ 第一步：下载 SKU 成本填写模板", data=buf.getvalue(), file_name="SKU_Cost_Template.xlsx")

if st.session_state.raw_df is not None:
    uploaded_cost = st.file_uploader("⬇️ 第二步：上传填好的成本表", type=["xlsx", "csv"], key="cost_up")
    if uploaded_cost:
        try:
            cdf = pd.read_excel(uploaded_cost) if str(uploaded_cost.name).endswith('xlsx') else pd.read_csv(uploaded_cost)
            cdf.columns = [str(c).strip() for c in cdf.columns]
            s_col = next((c for c in cdf.columns if 'sku' in c.lower()), None)
            c_col = next((c for c in cdf.columns if c == '产品和头程成本'), None)
            if s_col and c_col:
                costs = cdf[[s_col, c_col]].copy()
                costs[s_col] = costs[s_col].astype(str).str.strip()
                costs[c_col] = pd.to_numeric(costs[c_col], errors='coerce').fillna(0)
                st.session_state.sku_list = costs.rename(columns={s_col: 'SKU', c_col: '产品和头程成本'}).drop_duplicates(subset=['SKU'])
                st.success("✅ 成本匹配成功！")
            else:
                st.error("❌ 成本表列名错误。")
        except Exception as e: st.error(f"成本读取失败: {e}")

if st.session_state.sku_list is not None and 'type' in st.session_state.raw_df.columns:
    st.markdown("---")
    df, sc = st.session_state.raw_df, st.session_state.sku_list
    d_r = f"{df['date'].min().strftime('%Y/%m/%d')} - {df['date'].max().strftime('%Y/%m/%d')}" if not df['date'].isna().all() else "未知"
    
    SORT_PRIORITY = {'order': 1, 'refund': 2, 'cost of advertising': 3, 'subscription fee': 4, 'service fee': 5, 'fba inventory storage fee': 6, 'fba long-term storage fee': 7, 'fba inventory fee': 8, 'adjustment': 9, 'coupon redemption fee': 10, 'liquidations': 11}

    ts = df[~df['type'].isin(['transfer', 'debt'])].groupby('type').agg(q=('quantity','sum'), t=('total','sum')).reset_index()
    ts['sort_key'] = ts['type'].map(lambda x: SORT_PRIORITY.get(x, 99))
    ts = ts.sort_values(by=['sort_key', 'type']).reset_index(drop=True)
    
    df_t = pd.DataFrame([{"序号": i+1, "项目": TYPE_MAPPING.get(r['type'], r['type']), "TYPE": r['type'], "数量": r['q'] if r['q']!=0 else "-", "金额 (本币)": r['t']} for i, r in ts.iterrows()])
    df_t.loc[len(df_t)] = ["合计", "-", "Total", "-", df_t['金额 (本币)'].sum()]


    df_o = df[df['type'].isin(['order', 'refund'])]
    o_agg = df_o.groupby('sku').agg({'description':'first','product sales':'sum','net total':'sum','quantity':'sum'}).reset_index()
    res, trq, toq = [], 0, 0
    for _, r in o_agg.iterrows():
        sd = df_o[df_o['sku'] == r['sku']]
        o_o, r_o = sd[sd['type'] == 'order'], sd[sd['type'] == 'refund']
        oc, rq = o_o['quantity'].sum(), r_o['quantity'].sum()
        toq += oc; trq += rq
        res.append({'标题':r['description'],'SKU':r['sku'],'平均售价':(o_o['product sales'].sum()/oc) if oc>0 else 0,'销售额':o_o['product sales'].sum(),'售出回款':o_o['net total'].sum(),'退款金额':r_o['net total'].sum(),'退货数量':rq,'退货率':(rq/oc) if oc>0 else 0,'成交金额':r['net total'],'商品出库数量':oc,'商品成交数量':oc-rq})
    
    df_fo = pd.merge(pd.DataFrame(res), sc, on='SKU', how='left')
    df_fo['产品和头程成本'] = df_fo['产品和头程成本'].fillna(0)
    df_fo['订单成本'] = df_fo['商品成交数量'] * df_fo['产品和头程成本']
    df_fo['毛利'] = df_fo['成交金额'] - df_fo['订单成本']
    df_fo['单个利润'] = np.where(df_fo['商品成交数量']!=0, df_fo['毛利']/df_fo['商品成交数量'], 0)
    df_fo['毛利率'] = np.where(df_fo['销售额']!=0, df_fo['毛利']/df_fo['销售额'], 0)

    total_sales, total_receipts = df_fo['销售额'].sum(), df_fo['售出回款'].sum()
    total_refund_amt, total_refund_qty = df_fo['退款金额'].sum(), df_fo['退货数量'].sum()
    total_order_cost, total_gross_profit = df_fo['订单成本'].sum(), df_fo['毛利'].sum()
    t_tax = round((df['total']-df['net total']).sum(), 2)
    shop_p = df_t['金额 (本币)'].iloc[-1] - t_tax - total_order_cost
    g_m = shop_p / total_sales if total_sales != 0 else 0

    df_fo = df_fo[['标题','SKU','产品和头程成本','平均售价','销售额','售出回款','退款金额','退货数量','退货率','成交金额','商品出库数量','商品成交数量','订单成本','毛利','单个利润','毛利率']]
    df_fo.loc[len(df_fo)] = ['合计','-','-','-', total_sales, total_receipts, total_refund_amt, total_refund_qty, (trq/toq if toq else 0),'-','-','-', total_order_cost, total_gross_profit, '-', (total_gross_profit/total_sales if total_sales else 0)]
    
    buf_f = io.BytesIO()
    with pd.ExcelWriter(buf_f, engine='openpyxl') as w:
        df_fo.to_excel(w, index=False, sheet_name='Order', startrow=2)
        ws_o = w.sheets['Order']
        ws_o.merge_cells('A1:P1'); ws_o['A1']='订单统计表'; ws_o['A1'].font=Font(bold=True,size=16); ws_o['A1'].alignment=Alignment(horizontal='center')
        ws_o.column_dimensions['A'].width=40; ws_o.column_dimensions['B'].width=18
        for c in range(3,17): ws_o.column_dimensions[get_column_letter(c)].width=15
        for row in ws_o.iter_rows(min_row=3):
            for cell in row:
                cell.alignment=Alignment(horizontal='center')
                if cell.row==3 or cell.row==ws_o.max_row: cell.font=Font(bold=True); cell.fill=PatternFill(start_color="F2F2F2",fill_type="solid")
                if cell.column in [9,16] and cell.row>3: cell.number_format='0.00%'
                elif cell.column in [3,4,5,6,7,10,13,14,15] and cell.row>3: cell.number_format='0.00'
        
        df_t.to_excel(w, index=False, sheet_name='TOTAL', startrow=7)
        ws_t = w.sheets['TOTAL']
        ws_t.merge_cells('A1:E1'); ws_t['A1']='费用汇总'; ws_t['A1'].font=Font(bold=True,size=16); ws_t['A1'].alignment=Alignment(horizontal='center')
        

        stats = [("日期范围",d_r,"SKU数量",len(df_fo)-1), ("提现金额",abs(df[df['type']=='transfer']['total'].sum()),"存入金额",df[df['type']=='debt']['total'].sum()), ("总退货率",(trq/toq if toq else 0),"店铺利润",shop_p), ("产品和头程成本", total_order_cost, "毛利率", g_m), ("税金负债",t_tax,"","")]
        
        for r, (k1,v1,k2,v2) in enumerate(stats, 3):
            ws_t[f'B{r}']=k1; ws_t[f'C{r}']=v1; ws_t[f'D{r}']=k2; ws_t[f'E{r}']=v2
            ws_t[f'B{r}'].alignment = Alignment(horizontal='right', vertical='center')
            ws_t[f'C{r}'].alignment = Alignment(horizontal='center', vertical='center')
            ws_t[f'D{r}'].alignment = Alignment(horizontal='right', vertical='center')
            ws_t[f'E{r}'].alignment = Alignment(horizontal='center', vertical='center')
            

        ws_t['C4'].number_format = ws_t['E4'].number_format = ws_t['C6'].number_format = ws_t['C7'].number_format = '0.00'
        ws_t['C5'].number_format = ws_t['E6'].number_format = '0.00%'
        

        ws_t.column_dimensions['A'].width=10; ws_t.column_dimensions['B'].width=35; ws_t.column_dimensions['C'].width=30; ws_t.column_dimensions['D'].width=15; ws_t.column_dimensions['E'].width=18
        

        thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="F2F2F2", fill_type="solid")
        for row in ws_t.iter_rows(min_row=8):
            for cell in row:
                cell.alignment=Alignment(horizontal='center')
                if cell.row==8 or cell.row==ws_t.max_row: 
                    cell.font=Font(bold=True)
                    cell.border=thin_border
                if cell.row==8: 
                    cell.fill=header_fill
                if cell.column==5 and cell.row>8: 
                    cell.number_format='0.00'

        df_raw_exp = df.copy()
        if 'date' in df_raw_exp.columns: df_raw_exp['date'] = df_raw_exp['date'].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S') if pd.notnull(x) else x)
        df_raw_exp.to_excel(w, index=False, sheet_name='RawData')
        for c in range(1,len(df.columns)+1): w.sheets['RawData'].column_dimensions[get_column_letter(c)].width=20

    st.download_button("🚀 第三步：下载完整美化报表", data=buf_f.getvalue(), file_name=f"Amazon_Final_Report_{datetime.now().strftime('%Y%m%d')}.xlsx")
